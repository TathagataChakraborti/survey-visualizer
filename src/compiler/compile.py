from Levenshtein import distance as lev
from openpyxl import load_workbook
from typing import Dict, List, Optional
from sentence_transformers import SentenceTransformer
from schemas import Embedding, Taxonomy, Element, Paper, Network, Link
from sys import modules

import sklearn.manifold
import pdfplumber
import copy
import random
import argparse
import json
import glob
import os
import yaml  # type: ignore


def text_transform(text: str) -> str:
    text = text.replace("\n", " ").replace("-", "").lower()
    text = " ".join(text.split())
    return text


def generate_id(basis: int) -> int:
    return basis


def compute_parent(start_id: int, stop_id: int, data: List[List[Element]]) -> Optional[str]:
    if data:
        parent_candidates = data[-1]

        for parent in parent_candidates:
            if start_id >= parent.start and stop_id <= parent.stop:
                return parent.name

    return None


class Config2Data:
    def __init__(self, config: Dict):
        self.config = config
        self.cached_embeddings: List[Embedding] = []
        self.cached_taxonomy: Optional[Taxonomy] = None

    def get_taxonomy(self, config: Dict) -> List[Taxonomy]:
        out_data: List[Taxonomy] = list()

        for tab in config["tabs"]:
            print(f"Generating Taxonomy view for {tab['tab_name']}...")

            new_taxonomy = Taxonomy(
                name=tab["tab_name"],
                data=list(),
                taxonomy=list(),
            )

            path = tab["input_file"]["filename"]
            if tab["input_file"].get("relative"):
                path = os.path.abspath(
                    os.path.join(
                        os.path.dirname(os.path.realpath(__file__)), f"./{path}"
                    )
                )
            else:
                path = os.path.abspath(path)

            print(f"Looking for spreadshet at {path} ...")

            wb = load_workbook(path, data_only=True)
            sheet = wb[tab["input_file"]["active_worksheet"]]

            row_number = 0
            for row in sheet.iter_rows(max_row=sheet.max_row):
                row_number += 1

                if (
                    tab["taxonomy"]["rows"]["start"]
                    <= row_number
                    <= tab["taxonomy"]["rows"]["stop"]
                ):
                    if row_number in tab["taxonomy"]["rows"].get("exclude", []):
                        continue

                    row = row[tab["taxonomy"]["columns"]["start"] - 1: tab["taxonomy"]["columns"]["stop"]]
                    current_level = list()

                    new_item = None

                    for item in row:
                        if item.value:
                            if new_item:
                                current_level.append(new_item)
                                new_item = None

                            if not new_item:
                                name = item.value
                                start = tab["taxonomy"]["columns"]["start"] + row.index(
                                    item
                                )
                                stop = start

                                new_item = Element(
                                    name=name,
                                    expanded=True,
                                    parent=compute_parent(
                                        start, stop, new_taxonomy.taxonomy
                                    ),
                                    level=len(new_taxonomy.taxonomy) + 1,
                                    start=start,
                                    stop=stop,
                                )

                        if item.value is None and new_item is not None:
                            new_item.stop = tab["taxonomy"]["columns"][
                                "start"
                            ] + row.index(item)

                    if new_item:
                        current_level.append(new_item)

                    new_taxonomy.taxonomy.append(current_level)

                if (
                    tab["papers_list"]["rows"]["start"]
                    <= row_number
                    <= tab["papers_list"]["rows"]["stop"]
                ):
                    if row_number in tab["papers_list"]["rows"].get("exclude", []):
                        continue

                    existing_ids = [item.UID for item in new_taxonomy.data]
                    new_id = generate_id(row_number)

                    while new_id in existing_ids:
                        new_id = generate_id(row_number)

                    new_paper = Paper(
                        UID=new_id,
                        citations=list(),
                        selected=True,
                    )

                    key_map = tab["papers_list"]["key_map"]

                    for key in key_map:
                        if key_map[key] is not None:
                            setattr(new_paper, key, row[key_map[key]].value)
                        else:
                            setattr(new_paper, key, None)

                    if "year" in key_map:
                        new_paper.year = int(new_paper.year)

                    if "slug" not in key_map:
                        new_paper.slug = f"paper-{new_id}"

                    tags = list()
                    for item in row:
                        if item.value:
                            for taxonomic_level in new_taxonomy.taxonomy:
                                for taxonomic_item in taxonomic_level:
                                    if (
                                        taxonomic_item.start
                                        <= row.index(item) + 1
                                        <= taxonomic_item.stop
                                    ):
                                        new_tag_item = {
                                            "name": taxonomic_item.name,
                                            "parent": taxonomic_item.parent,
                                        }

                                        if new_tag_item not in tags:
                                            tags.append(new_tag_item)

                    new_paper.tags = tags
                    new_taxonomy.data.append(new_paper)

            if tab["papers_list"].get("shuffle_list", False):
                random.shuffle(new_taxonomy.data)

            out_data.append(new_taxonomy)

            if tab["tab_name"] == config["default_tab"]:
                self.cached_taxonomy = new_taxonomy

        return out_data

    def get_affinity(self, config: Dict) -> List[Embedding]:
        _ = config
        new_paper_list = list()

        for paper in self.cached_taxonomy.data:
            new_paper: Paper = copy.deepcopy(paper)

            new_paper.abstract = paper.abstract or paper.title
            new_paper.authors = paper.authors.replace(",", " | ")

            new_paper.keywords = paper.keywords or " | ".join(
                [tag["name"] for tag in paper.tags]
            )
            new_paper.sessions = paper.sessions or paper.venue
            new_paper_list.append(new_paper)

        # Referece: https://github.com/Mini-Conf/Mini-Conf/tree/master/scripts
        print("Generating Embeddings...")
        model = SentenceTransformer("allenai-specter")
        papers = [
            "[SEP]".join(
                [
                    getattr(paper, key) if type(getattr(paper, key)) == str and key != "UID" else ""
                    for key in paper.model_fields.keys()
                ]
            )
            for paper in new_paper_list
        ]
        embeddings = model.encode(papers, convert_to_tensor=True)

        print("Generating Transforms...")
        transform = sklearn.manifold.TSNE(n_components=2).fit_transform(embeddings.cpu().numpy())  # type: ignore

        out_data = []
        for i, row in enumerate(new_paper_list):
            position = transform[i].tolist()
            out_data.append(Embedding(UID=row.UID, x=position[0], y=position[1]))

        self.cached_embeddings = out_data
        return out_data

    def get_network(self, config: Dict) -> Network:
        network_data = Network(nodes=list(), links=list())

        for paper in self.cached_taxonomy.data:
            network_data.nodes.append(paper)

        path = config["files_directory"]["location"]

        if config["files_directory"].get("relative"):
            path = os.path.abspath(
                os.path.join(os.path.dirname(os.path.realpath(__file__)), f"./{path}")
            )
        else:
            path = os.path.abspath(path)

        match_threshold = config["match_threshold"]
        columns_per_page = [2, 1]

        # Reference: https://stackoverflow.com/a/69316574
        x0 = 0  # Distance of left side of character from left side of page.
        x1 = 0.5  # Distance of right side of character from left side of page.
        y0 = 0  # Distance of bottom of character from bottom of page.
        y1 = 1  # Distance of top of character from bottom of page.

        files_to_read = glob.glob(f"{path}/*.pdf")
        print(f"Looking for PDFs in {path} ...")

        for file in files_to_read:
            print(
                f"Parsing {files_to_read.index(file)+1}/{len(files_to_read)} papers..."
            )

            current_title = None
            current_id = None
            reference_map = []

            with pdfplumber.open(file) as pdf:
                for i, page in enumerate(pdf.pages):
                    start_references = False

                    for columns in columns_per_page:
                        text = page.extract_text()
                        transformed_text = text_transform(text)

                        if columns == 2:
                            try:
                                # Reference: https://stackoverflow.com/a/69316574
                                width = page.width
                                height = page.height

                                # Crop pages
                                left_bbox = (
                                    x0 * float(width),
                                    y0 * float(height),
                                    x1 * float(width),
                                    y1 * float(height),
                                )
                                page_crop = page.crop(bbox=left_bbox)
                                left_text = page_crop.extract_text()

                                left_bbox = (
                                    0.5 * float(width),
                                    y0 * float(height),
                                    1 * float(width),
                                    y1 * float(height),
                                )
                                page_crop = page.crop(bbox=left_bbox)
                                right_text = page_crop.extract_text()
                                page_context = "\n".join([left_text, right_text])
                                transformed_text = text_transform(page_context)

                            except ValueError as e:
                                print(e)
                                pass

                        if i == 0:
                            for paper in self.cached_taxonomy.data:
                                title_transform = text_transform(paper.title)

                                if title_transform in transformed_text:
                                    current_title = paper.title
                                    current_id = paper.UID
                                    break

                        if "references" in transformed_text:
                            transformed_text = transformed_text.split("references")[-1]
                            start_references = True

                        if start_references and current_title:
                            reference_list = transformed_text.split(".")

                            for reference in reference_list:
                                reference_map_item = None
                                current_match_score = match_threshold

                                for paper in self.cached_taxonomy.data:
                                    candidate_reference_title_transform = (
                                        text_transform(paper.title)
                                    )
                                    new_match_score = lev(
                                        reference, candidate_reference_title_transform
                                    ) / len(candidate_reference_title_transform)

                                    if new_match_score < current_match_score:
                                        current_match_score = new_match_score

                                        reference_map_item = {
                                            "reference_title": reference,
                                            "papers_title": paper.title,
                                            "reference_id": paper.UID,
                                            "match_score": current_match_score,
                                        }

                                if current_match_score < match_threshold:
                                    reference_map.append(reference_map_item)
            if current_title:
                for reference in reference_map:
                    new_link = Link(source=current_id, target=reference["reference_id"])
                    network_data.links.append(new_link)

        return network_data

    def get_insights(self, config: Dict) -> List[Embedding]:
        if self.cached_embeddings:
            return self.cached_embeddings
        else:
            return self.get_affinity(config)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Compile survey data.")
    parser.add_argument("--file", type=str, help="Path to YAML config file.")

    args = parser.parse_args()

    with open(args.file, "r") as stream:
        try:
            config_in = yaml.safe_load(stream)
        except yaml.YAMLError as exc:
            raise exc

    config_out = os.path.abspath(
        os.path.join(os.path.dirname(os.path.realpath(__file__)), f"./../config.json")
    )
    open(config_out, "w").write(json.dumps(config_in, indent=4))

    config_out_to_build = os.path.abspath(
        os.path.join(
            os.path.dirname(os.path.realpath(__file__)), f"./../../public/config.json"
        )
    )

    open(config_out_to_build, "w").write(
        json.dumps({"name": config_in["metadata"]["name"]}, indent=4)
    )

    reader_object = Config2Data(config_in)

    for view in config_in["views"]:
        view_name = view["name"]
        out_file = os.path.abspath(
            os.path.join(
                os.path.dirname(os.path.realpath(__file__)),
                f"./data/{view_name}.json",
            )
        )

        if not view["disabled"]:
            print(f"Generating {view_name} view...")
            function_name = f"get_{text_transform(view_name)}"

            module = modules[__name__]
            function = getattr(reader_object, function_name)
            output = function(view)

            if isinstance(output, List):
                output = [item.model_dump(warnings=False) for item in output]
            else:
                output = output.model_dump(warnings=False)
        else:
            output = dict()

        print(f"Writing {out_file}...")
        open(out_file, "w").write(json.dumps(output, indent=4))
